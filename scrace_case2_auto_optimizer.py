# -*- coding: utf-8 -*-
"""
SCRACE Case 2 - CKD Delivery Planning Auto Optimizer

Colab usage:
1) Upload this .py file and your Excel inputs, or paste all code into a Colab cell.
2) Set INPUT_FILE, WINDOW_FILE, OUTPUT_FILE below.
3) Run.

Main logic:
- Use ALL available production/inventory rows as supply. Feb rows are available inventory for March demand.
- Fulfill 100% dealer orders by model/SKU.
- FIFO is enforced at model/SKU level.
- Dispatch only after 4 WORKING HOURS after production output.
- Dispatch Monday-Saturday only, inside TMV shifts and optional dealer dispatch windows.
- Transport cost is calculated by TRIP/FULL CAPACITY, including empty slots.
- No multi-dealer combine shipment is applied by default; +300k/drop is therefore not charged.
"""

import math
import os
import re
import unicodedata
from difflib import SequenceMatcher
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =========================
# 0) CONFIG - EDIT HERE
# =========================
INPUT_FILE = "/content/Worksheet in SCRACE 26 Case study 13.05.xlsx"
WINDOW_FILE = "/content/SCRACE - KHUNG GIỜ ĐẠI LÍ.xlsx"  # optional; set None if not used
OUTPUT_FILE = "/content/SCRACE_case2_AUTO_FINAL.xlsx"

# If you run locally with files in /mnt/data, uncomment these lines:
# INPUT_FILE = "/mnt/data/Worksheet in SCRACE 26 Case study 13.05 (2).xlsx"
# WINDOW_FILE = "/mnt/data/SCRACE - KHUNG GIỜ ĐẠI LÍ.xlsx"
# OUTPUT_FILE = "/mnt/data/SCRACE_case2_AUTO_FINAL.xlsx"

CONFIG = {
    # scoring / trade-off target
    "target_inventory_days": 1.49,     # keep avg below 1.5 to get full inventory score
    "inventory_25pt_threshold": 1.50,
    "optimize_daily_cap": True,

    # constraints
    "fifo_level": "MODEL",            # business interpretation: FIFO per model/SKU
    "min_working_buffer_minutes": 240, # 4 working hours
    "no_sunday_dispatch": True,
    "arrival_must_be_in_same_month": True,
    "dispatch_must_be_in_company_shift": True,
    "dispatch_must_match_dealer_window": True,

    # shipment policy
    "allow_multi_dealer_combine": False,
    "extra_drop_fee_vnd": 300_000,

    # working shifts. The final shift crosses midnight.
    "company_shifts": [
        ("07:00", "11:10"),
        ("12:10", "15:55"),
        ("18:00", "22:20"),
        ("23:20", "02:55"),
    ],

    # Use date difference like the slide/formula: Ngày xuất bãi - Ngày xuất xưởng
    # Option: "date_diff" or "datetime_diff"
    "inventory_method": "date_diff",
}

MODEL_ORDER_FALLBACK = ["VE", "VK", "VG", "VLE0", "VLG0", "AE0", "AG0"]
CAPACITY_FALLBACK = {
    "Xe tải 1": 1,
    "Xe tải 2": 2,
    "Xe tải 4": 4,
    "Xe lồng 5": 5,
    "Xe lồng 6": 6,
    "Xe lồng 7": 7,
}

# =========================
# 1) BASIC HELPERS
# =========================

def norm_code(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip().upper()


def norm_text(x) -> str:
    if pd.isna(x):
        return ""
    return " ".join(str(x).strip().split())

def strip_vietnamese_accents(text: Any) -> str:
    """Lowercase + remove Vietnamese accents/combining marks for robust sheet matching."""
    if text is None:
        return ""
    txt = str(text).strip().lower().replace("đ", "d")
    txt = unicodedata.normalize("NFD", txt)
    txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
    txt = unicodedata.normalize("NFC", txt)
    return txt


def normalize_sheet_key(text: Any) -> str:
    txt = strip_vietnamese_accents(text)
    txt = re.sub(r"[^a-z0-9]+", "", txt)
    return txt


def list_excel_sheets(input_file: str) -> List[str]:
    try:
        return pd.ExcelFile(input_file).sheet_names
    except Exception as e:
        raise ValueError(f"Không đọc được file Excel: {input_file}. Chi tiết: {e}")


SHEET_ALIASES = {
    "orders": [
        "đơn đặt hàng", "don dat hang", "đại lý đặt hàng", "dai ly dat hang",
        "2.1 đại lý đặt hàng", "2.1 dai ly dat hang", "order", "orders",
        "dealer order", "dealer orders", "demand", "nhu cầu", "nhu cau",
    ],
    "production": [
        "kế hoạch sản xuất", "ke hoach san xuat", "2.2 kế hoạch sản xuất",
        "2.2 ke hoach san xuat", "production", "production plan", "plan",
        "khsx", "ke hoach", "sản xuất", "san xuat",
    ],
    "prices": [
        "bảng giá vận tải", "bang gia van tai", "2.3 bảng giá vận tải",
        "2.3 bang gia van tai", "price", "prices", "transport price",
        "transportation price", "freight", "bảng giá", "bang gia", "van tai",
    ],
}


def resolve_sheet_name(input_file: str, sheet_key: str, explicit_sheet: Optional[str] = None) -> str:
    """Resolve a sheet using exact, accent-insensitive and fuzzy matching.

    Raises a readable ValueError with the actual workbook sheet names instead of a raw pandas traceback.
    """
    sheets = list_excel_sheets(input_file)
    if not sheets:
        raise ValueError("Workbook không có sheet nào.")

    if explicit_sheet:
        explicit_norm = normalize_sheet_key(explicit_sheet)
        for sh in sheets:
            if normalize_sheet_key(sh) == explicit_norm:
                return sh
        raise ValueError(
            f"Sheet được chọn cho {sheet_key} là '{explicit_sheet}' nhưng không tồn tại trong file. "
            f"Các sheet hiện có: {sheets}"
        )

    aliases = SHEET_ALIASES.get(sheet_key, [])
    alias_norms = [normalize_sheet_key(a) for a in aliases]
    sheet_norms = {sh: normalize_sheet_key(sh) for sh in sheets}

    # 1) exact normalized match
    for sh, nsh in sheet_norms.items():
        if nsh in alias_norms:
            return sh

    # 2) contains match in either direction
    for sh, nsh in sheet_norms.items():
        for a in alias_norms:
            if a and (a in nsh or nsh in a):
                return sh

    # 3) fuzzy match with conservative threshold
    best = None
    best_score = 0.0
    for sh, nsh in sheet_norms.items():
        for a in alias_norms:
            score = SequenceMatcher(None, nsh, a).ratio()
            if score > best_score:
                best, best_score = sh, score
    if best and best_score >= 0.74:
        return best

    friendly = {
        "orders": "đơn đặt hàng",
        "production": "kế hoạch sản xuất",
        "prices": "bảng giá vận tải",
    }.get(sheet_key, sheet_key)
    raise ValueError(
        f"Không tìm thấy sheet '{friendly}'. Các sheet hiện có trong file: {sheets}. "
        "Hãy upload đúng worksheet chính hoặc chọn mapping sheet thủ công trong app."
    )


def parse_hhmm(s: str) -> time:
    s = str(s).strip()
    h, m = s.split(":")
    return time(int(h), int(m))


def combine_date_time(d, t) -> datetime:
    if isinstance(d, pd.Timestamp):
        d = d.to_pydatetime().date()
    elif isinstance(d, datetime):
        d = d.date()
    elif isinstance(d, date):
        pass
    else:
        d = pd.to_datetime(d).date()

    if isinstance(t, pd.Timestamp):
        t = t.to_pydatetime().time()
    elif isinstance(t, datetime):
        t = t.time()
    elif isinstance(t, time):
        pass
    elif pd.isna(t):
        t = time(0, 0)
    else:
        try:
            t = pd.to_datetime(t).time()
        except Exception:
            parts = str(t).split(":")
            t = time(int(parts[0]), int(parts[1]) if len(parts) > 1 else 0)
    return datetime.combine(d, t)


def month_bounds_from_production(prod_df: pd.DataFrame) -> Tuple[date, date]:
    # Use the dominant/latest month in the production plan.
    months = prod_df["prod_date"].dropna().dt.to_period("M")
    if len(months) == 0:
        raise ValueError("Cannot infer month from production plan.")
    target_period = months.mode().iloc[0]
    month_start = target_period.to_timestamp().date()
    next_month = (target_period + 1).to_timestamp().date()
    month_end = next_month - timedelta(days=1)
    return month_start, month_end


def is_sunday(dt: datetime) -> bool:
    return dt.weekday() == 6


def parse_window_text(s: Any) -> Tuple[time, time, bool]:
    """Return (start, end, is_24h). Window text examples: 'Đi 21h - 6h', 'đi 24/24'."""
    txt = norm_text(s).lower().replace("đ", "d")
    if not txt or "24/24" in txt or "24h" in txt:
        return time(0, 0), time(23, 59), True

    # normalize 21h30, 21:30, 21h
    nums = re.findall(r"(\d{1,2})\s*(?:h|:)(\d{0,2})", txt)
    if len(nums) >= 2:
        h1, m1 = nums[0]
        h2, m2 = nums[1]
        return time(int(h1), int(m1 or 0)), time(int(h2), int(m2 or 0)), False

    return time(0, 0), time(23, 59), True


def time_in_window(t: time, start: time, end: time, is_24h: bool = False) -> bool:
    if is_24h:
        return True
    if start <= end:
        return start <= t <= end
    return t >= start or t <= end


def datetime_in_daily_window(dt: datetime, start: time, end: time, is_24h: bool = False) -> bool:
    return time_in_window(dt.time(), start, end, is_24h)


def shift_intervals_for_date(d: date, shifts: List[Tuple[str, str]]) -> List[Tuple[datetime, datetime]]:
    intervals = []
    for st_s, en_s in shifts:
        st = parse_hhmm(st_s)
        en = parse_hhmm(en_s)
        a = datetime.combine(d, st)
        b = datetime.combine(d, en)
        if en <= st:
            b += timedelta(days=1)
        intervals.append((a, b))
    return intervals


def is_working_dt(dt: datetime, shifts: List[Tuple[str, str]], allow_sunday_buffer: bool = False) -> bool:
    # Count a cross-midnight shift by its shift start date. For dispatch, Sunday is blocked separately.
    for offset in [0, -1]:
        base_d = (dt + timedelta(days=offset)).date()
        if base_d.weekday() == 6 and not allow_sunday_buffer:
            continue
        if base_d.weekday() <= 5:
            for a, b in shift_intervals_for_date(base_d, shifts):
                if a <= dt < b:
                    return True
    return False


def next_working_dt(dt: datetime, shifts: List[Tuple[str, str]], allow_sunday_buffer: bool = False) -> datetime:
    if is_working_dt(dt, shifts, allow_sunday_buffer):
        return dt
    # Try next shift starts within 10 days.
    best = None
    for day_add in range(0, 15):
        d = (dt + timedelta(days=day_add)).date()
        if d.weekday() == 6 and not allow_sunday_buffer:
            continue
        if d.weekday() <= 5:
            for a, _ in shift_intervals_for_date(d, shifts):
                if a >= dt:
                    if best is None or a < best:
                        best = a
    if best is None:
        raise RuntimeError("Cannot find next working datetime.")
    return best


def add_working_minutes(start_dt: datetime, minutes: int, shifts: List[Tuple[str, str]]) -> datetime:
    """Add working minutes only inside company shifts. Off-shift time is paused."""
    remaining = int(minutes)
    cur = start_dt
    while remaining > 0:
        cur = next_working_dt(cur, shifts, allow_sunday_buffer=False)
        # find current interval end
        interval_end = None
        for offset in [0, -1]:
            base_d = (cur + timedelta(days=offset)).date()
            if base_d.weekday() <= 5:
                for a, b in shift_intervals_for_date(base_d, shifts):
                    if a <= cur < b:
                        interval_end = b
                        break
            if interval_end:
                break
        if interval_end is None:
            cur += timedelta(minutes=1)
            continue
        can_add = int((interval_end - cur).total_seconds() // 60)
        if remaining <= can_add:
            return cur + timedelta(minutes=remaining)
        remaining -= can_add
        cur = interval_end + timedelta(minutes=1)
    return cur


def is_dispatch_valid(dt: datetime, dealer: str, windows: Dict[str, Tuple[time, time, bool]], config: Dict) -> bool:
    if config.get("no_sunday_dispatch", True) and dt.weekday() == 6:
        return False
    if config.get("dispatch_must_be_in_company_shift", True):
        if not is_working_dt(dt, config["company_shifts"], allow_sunday_buffer=False):
            return False
    if config.get("dispatch_must_match_dealer_window", True):
        w = windows.get(dealer)
        if w is not None:
            if not datetime_in_daily_window(dt, *w):
                return False
    return True



def next_dealer_window_start(dt: datetime, w: Tuple[time, time, bool]) -> datetime:
    start, end, is_24h = w
    if is_24h or time_in_window(dt.time(), start, end, is_24h):
        return dt
    d = dt.date()
    # Non-crossing window: start <= end
    if start <= end:
        cand = datetime.combine(d, start)
        if dt <= cand:
            return cand
        return datetime.combine(d + timedelta(days=1), start)
    # Crossing midnight, e.g. 21:00-06:00. Invalid zone is after end and before start.
    cand = datetime.combine(d, start)
    if dt.time() < start and dt.time() > end:
        return cand
    return dt


def next_valid_dispatch_dt(dt: datetime, dealer: str, windows: Dict[str, Tuple[time, time, bool]], config: Dict) -> datetime:
    cur = dt.replace(second=0, microsecond=0)
    # Jump between company shifts and dealer-window starts instead of scanning minute-by-minute.
    for _ in range(500):
        if config.get("no_sunday_dispatch", True) and cur.weekday() == 6:
            cur = datetime.combine(cur.date() + timedelta(days=1), time(0, 0))
            continue
        if config.get("dispatch_must_be_in_company_shift", True) and not is_working_dt(cur, config["company_shifts"], False):
            nxt = next_working_dt(cur, config["company_shifts"], False)
            cur = max(nxt, cur + timedelta(minutes=1))
            continue
        if config.get("dispatch_must_match_dealer_window", True):
            w = windows.get(dealer)
            if w is not None and not datetime_in_daily_window(cur, *w):
                nxt = next_dealer_window_start(cur, w)
                cur = max(nxt, cur + timedelta(minutes=1))
                continue
        if is_dispatch_valid(cur, dealer, windows, config):
            return cur
        cur += timedelta(minutes=1)
    raise RuntimeError(f"Cannot find valid dispatch time for {dealer} after {dt}.")


def working_minutes_between(a: datetime, b: datetime, shifts: List[Tuple[str, str]]) -> int:
    if b <= a:
        return 0
    cur = a
    total = 0
    while cur < b:
        if is_working_dt(cur, shifts, allow_sunday_buffer=False):
            total += 1
        cur += timedelta(minutes=1)
    return total

# =========================
# 2) LOAD INPUTS
# =========================

def load_orders(input_file: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, List[str]]:
    resolved_sheet = resolve_sheet_name(input_file, "orders", sheet_name)
    raw = pd.read_excel(input_file, sheet_name=resolved_sheet, header=None)
    model_cols = []
    model_names = []
    for c in range(raw.shape[1]):
        val = raw.iloc[1, c]
        if isinstance(val, str) and norm_text(val).upper() in MODEL_ORDER_FALLBACK:
            model_cols.append(c)
            model_names.append(norm_text(val).upper())
    if not model_cols:
        model_cols = list(range(3, 10))
        model_names = MODEL_ORDER_FALLBACK[:len(model_cols)]

    rows = []
    for r in range(2, raw.shape[0]):
        dealer = norm_code(raw.iloc[r, 0])
        if not dealer:
            continue
        address = norm_text(raw.iloc[r, 1])
        lead = raw.iloc[r, 2]
        if pd.isna(lead):
            continue
        item = {"dealer": dealer, "address": address, "leadtime_hours": float(lead)}
        total = 0
        for col, m in zip(model_cols, model_names):
            qty = raw.iloc[r, col]
            qty = 0 if pd.isna(qty) else int(qty)
            item[m] = qty
            total += qty
        item["total_order"] = total
        rows.append(item)
    orders = pd.DataFrame(rows)
    return orders, model_names


def load_production(input_file: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    resolved_sheet = resolve_sheet_name(input_file, "production", sheet_name)
    raw = pd.read_excel(input_file, sheet_name=resolved_sheet, header=None)
    # Header row is Excel row 2 / zero-index 1
    rows = []
    for r in range(2, raw.shape[0]):
        stt = raw.iloc[r, 0]
        model = norm_text(raw.iloc[r, 1]).upper()
        if pd.isna(stt) or not model:
            continue
        prod_time = raw.iloc[r, 2]
        prod_date = raw.iloc[r, 3]
        try:
            prod_dt = combine_date_time(prod_date, prod_time)
        except Exception as e:
            raise ValueError(f"Cannot parse production datetime at Excel row {r+1}: {e}")
        rows.append({
            "excel_row": r + 1,
            "stt": int(stt),
            "model": model,
            "prod_date": pd.to_datetime(prod_dt.date()),
            "prod_time": prod_dt.time(),
            "prod_dt": prod_dt,
        })
    prod = pd.DataFrame(rows)
    prod = prod.sort_values(["model", "prod_dt", "stt"]).reset_index(drop=True)
    return prod


def load_prices(input_file: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, Dict[str, int]]:
    resolved_sheet = resolve_sheet_name(input_file, "prices", sheet_name)
    raw = pd.read_excel(input_file, sheet_name=resolved_sheet, header=None)

    cap_map = {}
    for r in range(1, min(raw.shape[0], 20)):
        mode = norm_text(raw.iloc[r, 6]) if raw.shape[1] > 7 else ""
        cap = raw.iloc[r, 7] if raw.shape[1] > 7 else None
        if mode and not pd.isna(cap):
            try:
                cap_map[mode] = int(cap)
            except Exception:
                pass
    if not cap_map:
        cap_map = CAPACITY_FALLBACK.copy()

    rows = []
    for r in range(1, raw.shape[0]):
        dealer = norm_code(raw.iloc[r, 0])
        mode = norm_text(raw.iloc[r, 1])
        unit_price = raw.iloc[r, 3] if raw.shape[1] > 3 else None
        if not dealer or not mode or pd.isna(unit_price):
            continue
        if mode not in cap_map:
            continue
        rows.append({
            "dealer": dealer,
            "mode": mode,
            "unit_price": float(unit_price),
            "capacity": int(cap_map[mode]),
            "trip_cost": float(unit_price) * int(cap_map[mode]),
        })
    prices = pd.DataFrame(rows).drop_duplicates(["dealer", "mode"]).reset_index(drop=True)
    return prices, cap_map


def load_windows(window_file: Optional[str]) -> Dict[str, Tuple[time, time, bool]]:
    windows = {}
    if not window_file or not os.path.exists(window_file):
        return windows
    raw = pd.read_excel(window_file, sheet_name=0, header=None)
    # Try to identify columns. In current file: A dealer, C window, data from row 3.
    for r in range(1, raw.shape[0]):
        dealer = norm_code(raw.iloc[r, 0]) if raw.shape[1] >= 1 else ""
        if not dealer:
            continue
        # choose first cell in the row containing 24/24 or 'Đi'
        wtxt = None
        for c in range(raw.shape[1]):
            val = raw.iloc[r, c]
            if isinstance(val, str) and ("24/24" in val.lower() or "đi" in val.lower() or "di" in val.lower()):
                wtxt = val
                break
        if wtxt is None and raw.shape[1] >= 3:
            wtxt = raw.iloc[r, 2]
        windows[dealer] = parse_window_text(wtxt)
    return windows

# =========================
# 3) VEHICLE/TRIP PLANNING
# =========================

def choose_vehicle_plan(dealer: str, demand: int, prices: pd.DataFrame) -> List[Dict[str, Any]]:
    """Unbounded DP: minimize cost for capacity >= demand. Tie: lower empty slots, fewer trips."""
    avail = prices[prices["dealer"] == dealer].copy()
    if avail.empty:
        raise ValueError(f"Dealer {dealer} has no transport price.")
    max_cap = int(avail["capacity"].max())
    upper = demand + max_cap
    # dp capacity -> (cost, trip_count, mode_list)
    dp = {0: (0.0, 0, [])}
    for total_cap in range(upper + 1):
        if total_cap not in dp:
            continue
        cur_cost, cur_n, cur_modes = dp[total_cap]
        for _, row in avail.iterrows():
            nc = total_cap + int(row["capacity"])
            if nc > upper:
                continue
            val = (cur_cost + float(row["trip_cost"]), cur_n + 1, cur_modes + [row["mode"]])
            if nc not in dp or (val[0], nc - demand if nc >= demand else 10**9, val[1]) < (dp[nc][0], nc - demand if nc >= demand else 10**9, dp[nc][1]):
                dp[nc] = val

    best = None
    best_cap = None
    for cap_total, val in dp.items():
        if cap_total >= demand:
            candidate_key = (val[0], cap_total - demand, val[1])
            if best is None or candidate_key < best[0]:
                best = (candidate_key, val)
                best_cap = cap_total
    if best is None:
        raise RuntimeError(f"Cannot build vehicle plan for {dealer}, demand {demand}.")

    modes = best[1][2]
    # Make trips more stable: large capacity first, then lower trip cost.
    mode_meta = avail.set_index("mode").to_dict("index")
    modes = sorted(modes, key=lambda m: (-mode_meta[m]["capacity"], mode_meta[m]["trip_cost"]))
    return [{"mode": m, **mode_meta[m]} for m in modes]


def build_trip_templates(orders: pd.DataFrame, prices: pd.DataFrame, models: List[str]) -> List[Dict[str, Any]]:
    trips = []
    trip_no = 1
    # Round-robin dealer order by total descending, then code.
    for _, row in orders.sort_values(["total_order", "dealer"], ascending=[False, True]).iterrows():
        dealer = row["dealer"]
        total = int(row["total_order"])
        if total <= 0:
            continue
        vehicle_plan = choose_vehicle_plan(dealer, total, prices)
        remaining = {m: int(row.get(m, 0) or 0) for m in models}
        # Fill each trip with a balanced mix from the dealer's remaining order.
        for v in vehicle_plan:
            cap = int(v["capacity"])
            load = {}
            loaded = 0
            while loaded < cap and sum(remaining.values()) > 0:
                # choose model with largest remaining, stable by model order
                m = max(models, key=lambda x: (remaining.get(x, 0), -models.index(x)))
                if remaining[m] <= 0:
                    break
                take = min(remaining[m], cap - loaded)
                load[m] = load.get(m, 0) + take
                remaining[m] -= take
                loaded += take
            trips.append({
                "trip_id": f"TR{trip_no:04d}",
                "dealer": dealer,
                "mode": v["mode"],
                "capacity": int(v["capacity"]),
                "unit_price": float(v["unit_price"]),
                "trip_cost": float(v["trip_cost"]),
                "loaded": int(loaded),
                "empty_slots": int(v["capacity"] - loaded),
                "load_by_model": load,
            })
            trip_no += 1
        if sum(remaining.values()) != 0:
            raise RuntimeError(f"Dealer {dealer} load assignment error: {remaining}")
    return trips


def round_robin_trip_sequence(trips: List[Dict[str, Any]], lead_map: Optional[Dict[str, float]] = None) -> List[Dict[str, Any]]:
    by_dealer: Dict[str, List[Dict[str, Any]]] = {}
    for t in trips:
        by_dealer.setdefault(t["dealer"], []).append(t)
    # Sort trips per dealer by capacity desc to ship larger loads first.
    for d in by_dealer:
        by_dealer[d] = sorted(by_dealer[d], key=lambda x: (-x["loaded"], x["trip_id"]))
    lead_map = lead_map or {}
    dealers = sorted(by_dealer.keys(), key=lambda d: (-float(lead_map.get(d, 0)), -sum(x["loaded"] for x in by_dealer[d]), d))
    out = []
    progress = True
    while progress:
        progress = False
        for d in dealers:
            if by_dealer[d]:
                out.append(by_dealer[d].pop(0))
                progress = True
    return out


def deadline_first_trip_sequence(trips: List[Dict[str, Any]], lead_map: Dict[str, float]) -> List[Dict[str, Any]]:
    """Prioritize long-leadtime dealers so distant dealers get earlier FIFO supply."""
    return sorted(trips, key=lambda t: (-float(lead_map.get(t["dealer"], 0)), t["dealer"], t["trip_id"]))

# =========================
# 4) SCHEDULING
# =========================

def assign_production_to_trips(trips: List[Dict[str, Any]], prod: pd.DataFrame, models: List[str]) -> Tuple[List[Dict[str, Any]], pd.DataFrame]:
    queues = {m: prod[prod["model"] == m].sort_values(["prod_dt", "stt"]).to_dict("records") for m in models}
    cursors = {m: 0 for m in models}
    car_records = []
    for trip in trips:
        trip_cars = []
        for m, qty in trip["load_by_model"].items():
            for _ in range(int(qty)):
                if cursors[m] >= len(queues[m]):
                    raise ValueError(f"Not enough production supply for model {m}.")
                car = queues[m][cursors[m]]
                cursors[m] += 1
                trip_cars.append(car)
        trip["cars"] = trip_cars
        trip["ready_dt"] = max(add_working_minutes(c["prod_dt"], CONFIG["min_working_buffer_minutes"], CONFIG["company_shifts"]) for c in trip_cars) if trip_cars else None
        for c in trip_cars:
            rec = c.copy()
            rec.update({"trip_id": trip["trip_id"], "dealer": trip["dealer"], "mode": trip["mode"]})
            car_records.append(rec)

    # Sanity: delivered count by model must not exceed demand. Remaining production is ending inventory.
    delivered = pd.DataFrame(car_records)
    return trips, delivered


def schedule_trips_with_cap(
    trips: List[Dict[str, Any]],
    orders: pd.DataFrame,
    windows: Dict[str, Tuple[time, time, bool]],
    daily_vehicle_cap: int,
    month_end: date,
    config: Dict,
) -> Tuple[List[Dict[str, Any]], pd.DataFrame, Dict[str, Any]]:
    lead_map = orders.set_index("dealer")["leadtime_hours"].to_dict()
    daily_usage: Dict[date, int] = {}
    prev_dispatch_by_model: Dict[str, datetime] = {}
    scheduled = []

    for trip in trips:
        earliest = trip["ready_dt"]
        # Enforce FIFO sequence per model: this trip cannot depart before previous trip carrying same model.
        for m, qty in trip["load_by_model"].items():
            if qty > 0 and m in prev_dispatch_by_model:
                earliest = max(earliest, prev_dispatch_by_model[m])

        # Find date/time satisfying daily cap and time windows.
        cur = earliest
        while True:
            cur = next_valid_dispatch_dt(cur, trip["dealer"], windows, config)
            d = cur.date()
            if config.get("no_sunday_dispatch", True) and d.weekday() == 6:
                cur = datetime.combine(d + timedelta(days=1), time(0, 0))
                continue
            if daily_usage.get(d, 0) + trip["loaded"] <= daily_vehicle_cap:
                break
            # move to next day earliest time
            cur = datetime.combine(d + timedelta(days=1), time(0, 0))
            if cur.date() > month_end:
                raise RuntimeError("Daily cap too low: cannot schedule within month.")

        dispatch_dt = cur
        lead_h = float(lead_map[trip["dealer"]])
        arrival_dt = dispatch_dt + timedelta(hours=lead_h)
        if config.get("arrival_must_be_in_same_month", True) and arrival_dt.date() > month_end:
            raise RuntimeError(f"Arrival after month-end for trip {trip['trip_id']} ({trip['dealer']}).")
        daily_usage[dispatch_dt.date()] = daily_usage.get(dispatch_dt.date(), 0) + trip["loaded"]

        trip2 = trip.copy()
        trip2.update({
            "dispatch_dt": dispatch_dt,
            "arrival_dt": arrival_dt,
            "leadtime_hours": lead_h,
        })
        scheduled.append(trip2)
        for m, qty in trip2["load_by_model"].items():
            if qty > 0:
                prev_dispatch_by_model[m] = dispatch_dt

    car_rows = []
    for trip in scheduled:
        cost_per_loaded_car = trip["trip_cost"] / max(1, trip["loaded"])
        for c in trip["cars"]:
            if config["inventory_method"] == "date_diff":
                inv_days = (trip["dispatch_dt"].date() - c["prod_dt"].date()).days
            else:
                inv_days = (trip["dispatch_dt"] - c["prod_dt"]).total_seconds() / 86400
            car_rows.append({
                **c,
                "dealer": trip["dealer"],
                "mode": trip["mode"],
                "trip_id": trip["trip_id"],
                "dispatch_dt": trip["dispatch_dt"],
                "arrival_dt": trip["arrival_dt"],
                "inventory_days": inv_days,
                "allocated_cost": cost_per_loaded_car,
                "trip_cost": trip["trip_cost"],
                "trip_loaded": trip["loaded"],
                "trip_capacity": trip["capacity"],
            })
    cars = pd.DataFrame(car_rows)
    metrics = compute_metrics(cars, scheduled, orders, month_end, config)
    return scheduled, cars, metrics


def working_days_in_month(month_start: date, month_end: date) -> List[date]:
    out = []
    d = month_start
    while d <= month_end:
        if d.weekday() <= 5:
            out.append(d)
        d += timedelta(days=1)
    return out


def optimize_schedule(trips: List[Dict[str, Any]], orders: pd.DataFrame, windows: Dict[str, Tuple[time, time, bool]], month_start: date, month_end: date, config: Dict):
    total_loaded = sum(t["loaded"] for t in trips)
    max_trip_loaded = max(t["loaded"] for t in trips)
    workdays = working_days_in_month(month_start, month_end)
    avg_cap = math.ceil(total_loaded / max(1, len(workdays)))
    low = max(max_trip_loaded, avg_cap)
    high = total_loaded

    scenario_rows = []
    best = None

    if not config.get("optimize_daily_cap", True):
        cap_candidates = [high]
    else:
        # Search from smoothest feasible to ASAP. Step 1 for accuracy at case-study size.
        cap_candidates = list(range(low, min(high, max(low + 250, 350)) + 1, 5))
        if high not in cap_candidates:
            cap_candidates.append(high)

    for cap in cap_candidates:
        import copy
        trial_trips = copy.deepcopy(trips)
        try:
            scheduled, cars, metrics = schedule_trips_with_cap(trial_trips, orders, windows, cap, month_end, config)
            ok = metrics["avg_inventory_days"] < config["inventory_25pt_threshold"] and metrics["arrival_after_month_count"] == 0
            scenario_rows.append({"daily_vehicle_cap": cap, **metrics, "feasible": True, "selected_candidate": ok})
            if ok and best is None:
                best = (scheduled, cars, metrics, cap)
                break
        except Exception as e:
            scenario_rows.append({"daily_vehicle_cap": cap, "feasible": False, "error": str(e)[:250]})
            continue

    if best is None:
        # Fallback ASAP/high cap if target inventory cannot be met, but still generate a plan if feasible.
        import copy
        scheduled, cars, metrics = schedule_trips_with_cap(copy.deepcopy(trips), orders, windows, high, month_end, config)
        best = (scheduled, cars, metrics, high)
        scenario_rows.append({"daily_vehicle_cap": high, **metrics, "feasible": True, "selected_candidate": True})

    scenario_df = pd.DataFrame(scenario_rows)
    return best[0], best[1], best[2], best[3], scenario_df

# =========================
# 5) METRICS + CHECKS
# =========================

def ratio_max_min_avg(values: List[float]) -> float:
    vals = [float(v) for v in values if not pd.isna(v)]
    if not vals:
        return 0.0
    avg = sum(vals) / len(vals)
    if avg == 0:
        return 0.0
    return (max(vals) - min(vals)) / avg


def compute_heijunka(cars: pd.DataFrame, trips: List[Dict[str, Any]], orders: pd.DataFrame) -> Dict[str, float]:
    if cars.empty:
        return {"H1_daily_total_ratio": 0, "H2_mode_daily_ratio": 0, "H3_dealer_trip_ratio": 0, "H4_trip_mix_ratio": 0, "heijunka_worst_ratio": 0}
    c = cars.copy()
    c["dispatch_date"] = c["dispatch_dt"].dt.date
    daily = c.groupby("dispatch_date").size().tolist()
    h1 = ratio_max_min_avg(daily)

    h2_parts = []
    for mode, g in c.groupby("mode"):
        h2_parts.append(ratio_max_min_avg(g.groupby("dispatch_date").size().tolist()))
    h2 = max(h2_parts) if h2_parts else 0

    trip_df = pd.DataFrame([{
        "trip_id": t["trip_id"], "dealer": t["dealer"], "loaded": t["loaded"], "mode": t["mode"]
    } for t in trips])
    dealer_trip_counts = trip_df.groupby("dealer").size().tolist()
    # include dealers with zero? orders all >0 in case study; keep active dealers only.
    h3 = ratio_max_min_avg(dealer_trip_counts)

    h4_parts = []
    for t in trips:
        mix = [q for q in t["load_by_model"].values() if q > 0]
        if len(mix) >= 2:
            h4_parts.append(ratio_max_min_avg(mix))
        elif len(mix) == 1:
            h4_parts.append(0.0)
    h4 = max(h4_parts) if h4_parts else 0.0
    return {
        "H1_daily_total_ratio": h1,
        "H2_mode_daily_ratio": h2,
        "H3_dealer_trip_ratio": h3,
        "H4_trip_mix_ratio": h4,
        "heijunka_worst_ratio": max(h1, h2, h3, h4),
    }


def score_inventory(avg_inv: float) -> int:
    if avg_inv < 1.5:
        return 25
    if avg_inv < 2.0:
        return 20
    if avg_inv < 2.5:
        return 15
    if avg_inv < 3.0:
        return 10
    return 0


def score_heijunka(worst_ratio: float) -> int:
    if worst_ratio < 0.10:
        return 15
    if worst_ratio < 0.20:
        return 10
    if worst_ratio < 0.30:
        return 5
    return 0


def compute_metrics(cars: pd.DataFrame, trips: List[Dict[str, Any]], orders: pd.DataFrame, month_end: date, config: Dict) -> Dict[str, Any]:
    total_order = int(orders["total_order"].sum())
    total_delivered = len(cars)
    total_trip_cost = sum(float(t["trip_cost"]) for t in trips)
    total_slots = sum(int(t["capacity"]) for t in trips)
    total_loaded = sum(int(t["loaded"]) for t in trips)
    empty_slots = total_slots - total_loaded
    avg_inv = float(cars["inventory_days"].mean()) if not cars.empty else 0.0
    arrival_after = int((cars["arrival_dt"].dt.date > month_end).sum()) if not cars.empty else 0
    sunday_dispatch = int((cars["dispatch_dt"].dt.weekday == 6).sum()) if not cars.empty else 0

    h = compute_heijunka(cars, trips, orders)
    return {
        "total_order": total_order,
        "total_delivered": total_delivered,
        "dealer_count": int((orders["total_order"] > 0).sum()),
        "trip_count": len(trips),
        "total_trip_cost": total_trip_cost,
        "total_slots_paid": total_slots,
        "total_loaded": total_loaded,
        "empty_slots_paid": empty_slots,
        "slot_utilization": total_loaded / total_slots if total_slots else 0,
        "avg_inventory_days": avg_inv,
        "inventory_score": score_inventory(avg_inv),
        "arrival_after_month_count": arrival_after,
        "sunday_dispatch_vehicle_count": sunday_dispatch,
        **h,
        "heijunka_score": score_heijunka(h["heijunka_worst_ratio"]),
        "fifo_score_if_model_level": 5,
    }


def build_checks(cars: pd.DataFrame, prod: pd.DataFrame, orders: pd.DataFrame, prices: pd.DataFrame, month_end: date) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    # Dealer-model coverage
    model_cols = [c for c in orders.columns if c in MODEL_ORDER_FALLBACK]
    delivered_pivot = cars.pivot_table(index="dealer", columns="model", values="stt", aggfunc="count", fill_value=0)
    rows = []
    for _, o in orders.iterrows():
        for m in model_cols:
            order_qty = int(o[m])
            delivered_qty = int(delivered_pivot.loc[o["dealer"], m]) if o["dealer"] in delivered_pivot.index and m in delivered_pivot.columns else 0
            if order_qty or delivered_qty:
                rows.append({"dealer": o["dealer"], "model": m, "order_qty": order_qty, "delivered_qty": delivered_qty, "diff": delivered_qty - order_qty})
    coverage = pd.DataFrame(rows)

    # FIFO by model check: sequence of delivered vehicles by dispatch should have nondecreasing production datetime within each model.
    fifo_rows = []
    violations = []
    for m, g in cars.sort_values(["dispatch_dt", "prod_dt", "stt"]).groupby("model"):
        prev_prod = None
        prev_stt = None
        v = 0
        for _, r in g.iterrows():
            if prev_prod is not None and r["prod_dt"] < prev_prod:
                v += 1
                violations.append({"model": m, "stt": r["stt"], "prod_dt": r["prod_dt"], "dispatch_dt": r["dispatch_dt"], "prev_prod_dt": prev_prod, "prev_stt": prev_stt})
            prev_prod = r["prod_dt"]
            prev_stt = r["stt"]
        fifo_rows.append({"model": m, "delivered": len(g), "fifo_model_violations": v})
    fifo_check = pd.DataFrame(fifo_rows)

    # Trip price mode coverage
    used_pairs = cars[["dealer", "mode"]].drop_duplicates()
    price_pairs = set(tuple(x) for x in prices[["dealer", "mode"]].drop_duplicates().values.tolist())
    used_pairs["has_price"] = used_pairs.apply(lambda r: (r["dealer"], r["mode"]) in price_pairs, axis=1)
    return coverage, fifo_check, used_pairs

# =========================
# 6) EXPORT OUTPUT WORKBOOK
# =========================

def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(start_row, j, col)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=start_col):
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            elif isinstance(val, np.generic):
                val = val.item()
            ws.cell(i, j, val)


def style_sheet(ws, freeze: str = "A2"):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx in range(1, min(ws.max_column, 30) + 1):
        max_len = 8
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(ws.max_row, 200)):
            for c in cell:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value))[:60] if False else min(len(str(c.value)), 45))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 36)


def get_openpyxl_sheet_by_resolved_name(wb, desired_sheet: str) -> str:
    """Return an actual openpyxl worksheet name using robust normalized matching."""
    if desired_sheet in wb.sheetnames:
        return desired_sheet
    desired_norm = normalize_sheet_key(desired_sheet)
    for sh in wb.sheetnames:
        if normalize_sheet_key(sh) == desired_norm:
            return sh
    for sh in wb.sheetnames:
        nsh = normalize_sheet_key(sh)
        if desired_norm and (desired_norm in nsh or nsh in desired_norm):
            return sh
    raise ValueError(
        f"Không tìm thấy sheet kế hoạch sản xuất để ghi output. "
        f"Sheet mong muốn: '{desired_sheet}'. Các sheet hiện có: {wb.sheetnames}"
    )


def export_workbook(input_file: str, output_file: str, cars: pd.DataFrame, trips: List[Dict[str, Any]], orders: pd.DataFrame, prod: pd.DataFrame, prices: pd.DataFrame, scenario_df: pd.DataFrame, metrics: Dict[str, Any], month_end: date, production_sheet: Optional[str] = None):
    wb = load_workbook(input_file)
    resolved_prod_sheet = resolve_sheet_name(input_file, "production", production_sheet)
    actual_ws_name = get_openpyxl_sheet_by_resolved_name(wb, resolved_prod_sheet)
    ws = wb[actual_ws_name]

    # Map STT to output records.
    car_by_stt = {int(r["stt"]): r for _, r in cars.iterrows()}

    for row in range(3, ws.max_row + 1):
        stt = ws.cell(row, 1).value
        if stt is None:
            continue
        try:
            stt_i = int(stt)
        except Exception:
            continue
        r = car_by_stt.get(stt_i)
        if r is None:
            # ending inventory / unassigned
            for col in range(5, 13):
                if col != 11:  # keep formula if wanted? clear all output fields except inventory formula can remain blank
                    ws.cell(row, col, None)
            continue
        dispatch_dt = r["dispatch_dt"].to_pydatetime() if isinstance(r["dispatch_dt"], pd.Timestamp) else r["dispatch_dt"]
        arrival_dt = r["arrival_dt"].to_pydatetime() if isinstance(r["arrival_dt"], pd.Timestamp) else r["arrival_dt"]
        ws.cell(row, 5, r["dealer"])
        ws.cell(row, 6, dispatch_dt.date())
        ws.cell(row, 7, dispatch_dt.time())
        ws.cell(row, 8, arrival_dt.date())
        ws.cell(row, 9, arrival_dt.time())
        ws.cell(row, 10, r["mode"])
        ws.cell(row, 11, float(r["inventory_days"]))
        ws.cell(row, 12, float(r["allocated_cost"]))

    # formats
    for row in range(3, ws.max_row + 1):
        ws.cell(row, 6).number_format = "d-mmm"
        ws.cell(row, 7).number_format = "hh:mm"
        ws.cell(row, 8).number_format = "d-mmm"
        ws.cell(row, 9).number_format = "hh:mm"
        ws.cell(row, 11).number_format = "0.00"
        ws.cell(row, 12).number_format = '#,##0'

    # Remove/recreate output sheets
    for s in ["Score_Dashboard", "Trip_Summary", "Dealer_Coverage", "FIFO_Check", "Price_Mode_Check", "Heijunka_Check", "Scenario_Tradeoff", "Algorithm_Notes", "Combine_Shipments_Module"]:
        if s in wb.sheetnames:
            del wb[s]

    # Trip summary
    trip_rows = []
    for t in trips:
        trip_rows.append({
            "trip_id": t["trip_id"],
            "dealer": t["dealer"],
            "mode": t["mode"],
            "capacity": t["capacity"],
            "loaded": t["loaded"],
            "empty_slots": t["empty_slots"],
            "unit_price_per_slot": t["unit_price"],
            "trip_cost_full_capacity": t["trip_cost"],
            "dispatch_dt": t["dispatch_dt"],
            "arrival_dt": t["arrival_dt"],
            "load_by_model": str(t["load_by_model"]),
            "multi_drop_applied": False,
            "extra_drop_fee": 0,
        })
    trip_df = pd.DataFrame(trip_rows)

    coverage, fifo_check, price_mode_check = build_checks(cars, prod, orders, prices, month_end)

    dashboard = pd.DataFrame([
        ["Total order vehicles", metrics["total_order"]],
        ["Total delivered vehicles", metrics["total_delivered"]],
        ["Dealer count", metrics["dealer_count"]],
        ["Trip count", metrics["trip_count"]],
        ["Total paid slots", metrics["total_slots_paid"]],
        ["Loaded slots", metrics["total_loaded"]],
        ["Empty slots paid", metrics["empty_slots_paid"]],
        ["Slot utilization", metrics["slot_utilization"]],
        ["Total transport cost VND", metrics["total_trip_cost"]],
        ["Avg inventory days", metrics["avg_inventory_days"]],
        ["Inventory score", metrics["inventory_score"]],
        ["FIFO score if model-level", metrics["fifo_score_if_model_level"]],
        ["H1 daily total ratio", metrics["H1_daily_total_ratio"]],
        ["H2 mode daily ratio", metrics["H2_mode_daily_ratio"]],
        ["H3 dealer trip ratio", metrics["H3_dealer_trip_ratio"]],
        ["H4 trip mix ratio", metrics["H4_trip_mix_ratio"]],
        ["Heijunka worst ratio", metrics["heijunka_worst_ratio"]],
        ["Heijunka score", metrics["heijunka_score"]],
        ["Arrival after month count", metrics["arrival_after_month_count"]],
        ["Sunday dispatch vehicle count", metrics["sunday_dispatch_vehicle_count"]],
        ["Multi-dealer combine applied", "NO"],
        ["Extra drop fee charged", 0],
    ], columns=["KPI", "Value"])

    heijunka = pd.DataFrame([
        ["H1 - Total vehicles dispatched per day", metrics["H1_daily_total_ratio"], "(max-min)/average"],
        ["H2 - Vehicles by mode per day", metrics["H2_mode_daily_ratio"], "worst mode ratio"],
        ["H3 - Trips per dealer in month", metrics["H3_dealer_trip_ratio"], "(max-min)/average"],
        ["H4 - Model mix within trip", metrics["H4_trip_mix_ratio"], "worst trip mix ratio"],
        ["Worst of H1-H4", metrics["heijunka_worst_ratio"], "used for scoring"],
    ], columns=["Heijunka_metric", "ratio", "formula_note"])

    notes = pd.DataFrame([
        ["FIFO", "Applied per model/SKU. Each model's earliest production rows are assigned and dispatched in nondecreasing production order. This fits dealer orders by model."],
        ["February inventory", "Rows with February production dates are treated as available supply/inventory for March demand."],
        ["4-hour buffer", "Dispatch is allowed only after 240 minutes of actual company working time after production output. Off-shift time is paused."],
        ["Cost", "Cost is calculated per trip/full capacity = capacity x quoted unit price. Empty slots are still paid and allocated back to loaded vehicles."],
        ["Transport mode", "Only modes existing in the dealer price table are allowed. No synthetic price is created."],
        ["Multi-drop combine", "Not applied in this final plan. Therefore +300,000 VND/drop is not charged. Add Google Maps distance matrix before applying multi-dealer combine."],
        ["Optimization", "The system searches daily vehicle cap scenarios and selects the smoothest feasible plan with average inventory below the target threshold."],
    ], columns=["Topic", "Explanation"])

    combine_module = pd.DataFrame([
        ["Current final plan", "Direct/single-dealer trip only; vehicles are grouped by same dealer, mode, and dispatch time."],
        ["When to apply multi-drop", "Only if Google Maps/distance matrix proves dealers are on the same route and saving after extra drop fee is positive."],
        ["Cost formula", "combined_trip_cost = capacity * unit_price_of_farthest/base_dealer + 300000 * extra_drop_count"],
        ["Safety checks", "Do not violate capacity, FIFO, 4 working-hour buffer, arrival month deadline, and dealer dispatch window."],
    ], columns=["Item", "Rule"])

    sheet_data = {
        "Score_Dashboard": dashboard,
        "Trip_Summary": trip_df,
        "Dealer_Coverage": coverage,
        "FIFO_Check": fifo_check,
        "Price_Mode_Check": price_mode_check,
        "Heijunka_Check": heijunka,
        "Scenario_Tradeoff": scenario_df,
        "Algorithm_Notes": notes,
        "Combine_Shipments_Module": combine_module,
    }

    for name, df in sheet_data.items():
        ws2 = wb.create_sheet(name)
        write_df(ws2, df)
        style_sheet(ws2)
        if name == "Score_Dashboard":
            ws2.column_dimensions["A"].width = 35
            ws2.column_dimensions["B"].width = 28
            for r in range(2, ws2.max_row + 1):
                if isinstance(ws2.cell(r, 2).value, float):
                    ws2.cell(r, 2).number_format = "#,##0.0000"

    # Style main plan minimally
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A3"
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["J"].width = 18

    wb.save(output_file)

# =========================
# 7) MAIN RUNNER
# =========================

def run_optimizer(input_file: str = INPUT_FILE, window_file: Optional[str] = WINDOW_FILE, output_file: str = OUTPUT_FILE, config: Dict = CONFIG, sheet_map: Optional[Dict[str, str]] = None):
    print("Loading input files...")
    sheet_map = sheet_map or {}
    orders, models = load_orders(input_file, sheet_map.get("orders"))
    prod = load_production(input_file, sheet_map.get("production"))
    prices, cap_map = load_prices(input_file, sheet_map.get("prices"))
    windows = load_windows(window_file)
    month_start, month_end = month_bounds_from_production(prod)

    print(f"Models: {models}")
    print(f"Dealers: {len(orders)}, total order: {orders['total_order'].sum():,.0f}")
    print(f"Production rows: {len(prod):,.0f}. Target month: {month_start} to {month_end}")
    print(f"Price rows: {len(prices):,.0f}. Dealer windows loaded: {len(windows):,.0f}")

    # Validate supply by model.
    order_by_model = orders[models].sum().astype(int).to_dict()
    supply_by_model = prod.groupby("model").size().to_dict()
    for m in models:
        if supply_by_model.get(m, 0) < order_by_model.get(m, 0):
            raise ValueError(f"Not enough supply for model {m}: supply={supply_by_model.get(m,0)}, demand={order_by_model.get(m,0)}")

    missing_price_dealers = sorted(set(orders.loc[orders["total_order"] > 0, "dealer"]) - set(prices["dealer"]))
    if missing_price_dealers:
        raise ValueError(f"Dealers missing price table: {missing_price_dealers}")

    print("Building cost-minimized single-dealer trip templates...")
    trip_templates = build_trip_templates(orders, prices, models)
    lead_map = orders.set_index("dealer")["leadtime_hours"].to_dict()
    # Deadline-first is safer for month-end feasibility; round-robin can be used later for stronger H3.
    trip_sequence = deadline_first_trip_sequence(trip_templates, lead_map)

    print("Assigning production rows to trips by model-level FIFO...")
    trips_with_cars, _ = assign_production_to_trips(trip_sequence, prod, models)

    # Important scheduling improvement:
    # After FIFO assignment, dispatch the trips whose cars are ready earlier first.
    # This prevents early inventory rows from being delayed behind later-production trips
    # and makes the plan much closer to the intended low-inventory trade-off.
    lead_map_for_sort = orders.set_index("dealer")["leadtime_hours"].to_dict()
    trips_with_cars = sorted(
        trips_with_cars,
        key=lambda t: (t["ready_dt"], -float(lead_map_for_sort.get(t["dealer"], 0)), t["dealer"], t["trip_id"])
    )

    print("Optimizing dispatch schedule with daily cap search...")
    scheduled_trips, cars, metrics, selected_cap, scenario_df = optimize_schedule(trips_with_cars, orders, windows, month_start, month_end, config)
    print(f"Selected daily vehicle cap: {selected_cap}")
    print("Key metrics:")
    for k in ["total_delivered", "trip_count", "total_trip_cost", "avg_inventory_days", "inventory_score", "heijunka_worst_ratio", "heijunka_score", "empty_slots_paid", "slot_utilization"]:
        print(f"  {k}: {metrics[k]}")

    print(f"Exporting output workbook: {output_file}")
    export_workbook(input_file, output_file, cars, scheduled_trips, orders, prod, prices, scenario_df, metrics, month_end, production_sheet=sheet_map.get("production"))
    print("DONE")
    return output_file, metrics


if __name__ == "__main__":
    run_optimizer()
